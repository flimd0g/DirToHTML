import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment
import os
import re

CONFIG_FILE = "config.txt"

def load_configuration():
    if os.path.isfile(CONFIG_FILE):
        with open(CONFIG_FILE, 'r') as file:
            excel_file_path = file.readline().strip()
            if excel_file_path:
                excel_file_entry.delete(0, tk.END)
                excel_file_entry.insert(0, excel_file_path)
                return excel_file_path
    return None

def save_configuration(excel_file_path):
    with open(CONFIG_FILE, 'w') as file:
        file.write(excel_file_path)

def select_excel_file():
    excel_file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if excel_file_path:
        excel_file_entry.delete(0, tk.END)
        excel_file_entry.insert(0, excel_file_path)
        save_configuration(excel_file_path)

def select_directory():
    directory_path = filedialog.askdirectory()
    job_number = job_number_entry.get()
    excel_path = excel_file_entry.get()
    if directory_path and job_number and excel_path:
        try:
            process_directory(directory_path, job_number, excel_path)
            messagebox.showinfo("Success", "Excel file updated successfully.")
            root.destroy()  # Close the GUI after success
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

def parse_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    keys_to_extract = [
        'ACM hardware class', 'ACM version', 'ACM diagnosis version', 'ACM VIN', 'ACM serial number',
        'ACM hardware part number', 'ACM certification', 'ACM hardware version'
    ]

    extracted_values = {key: None for key in keys_to_extract}

    rows = soup.find_all('tr')
    for row in rows:
        cells = row.find_all('td')
        if len(cells) == 2:
            key = cells[0].get_text(strip=True)
            value = cells[1].get_text(strip=True)
            if key == 'ACM diagnosis version':
                value = value.lstrip('0')
            if key in extracted_values:
                extracted_values[key] = value

    for key, value in extracted_values.items():
        print(f"{key}: {value}")

    return extracted_values

def update_excel(extracted_values, job_number, excel_path):
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header_mapping = {
        'ACM hardware class': 'Hardware Class',
        'ACM version': 'Version',
        'ACM diagnosis version': 'Diagnosis Version',
        'ACM VIN': 'Vin',
        'ACM serial number': 'Serial Number',
        'ACM hardware part number': 'Part Number',
        'ACM certification': 'Certification',
        'ACM hardware version': 'Hardware Version',
        'Job number': 'Fixably No.'
    }

    header_row_index = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        headers = {cell.value: cell.column for cell in row if cell.value}
        print(f"Headers found in row {row[0].row}: {headers}")
        if set(header_mapping.values()).issubset(headers.keys()):
            header_row_index = row[0].row
            break

    if not header_row_index:
        raise ValueError("Header row not found in the Excel sheet")

    headers = {cell.value: cell.column for cell in ws[header_row_index]}
    print(f"Headers and their columns: {headers}")

    for key in extracted_values.keys():
        if header_mapping[key] not in headers:
            raise ValueError(f"Column for '{key}' not found in the Excel sheet")

    extracted_values['Job number'] = job_number

    target_row = None
    for row in ws.iter_rows(min_row=header_row_index + 1):
        id_cell = row[0]
        if id_cell.value is not None:
            if all(cell.value is None for cell in row if cell.column != 1):
                target_row = id_cell.row
                break

    if target_row is None:
        print("No suitable row found for updating")
        raise ValueError("No suitable row found for updating")

    for key, value in extracted_values.items():
        cell = ws.cell(row=target_row, column=headers[header_mapping[key]], value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(excel_path)
    wb.close()

def process_file(file_path, job_number, excel_path, existing_job_numbers):
    extracted_values = parse_html(file_path)
    acm_vin = extracted_values.get('ACM VIN', '')

    # Check if the job number already exists
    if acm_vin in existing_job_numbers:
        print(f"Skipping file {file_path} as it is already processed")
        return

    update_excel(extracted_values, job_number, excel_path)
    existing_job_numbers.add(acm_vin)

def process_directory(directory_path, job_number, excel_path):
    existing_job_numbers = set()
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Collect existing job numbers from the Excel sheet
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        acm_vin = row[ws['Vin'].column - 1].value
        if acm_vin:
            existing_job_numbers.add(acm_vin)

    wb.close()

    for file_name in os.listdir(directory_path):
        if file_name.endswith('.html'):
            file_path = os.path.join(directory_path, file_name)
            process_file(file_path, job_number, excel_path, existing_job_numbers)

# GUI Setup
root = tk.Tk()
root.title("HTML to Excel")

# Create a Notebook (tabbed interface)
notebook = ttk.Notebook(root)
notebook.pack(padx=10, pady=10, expand=True, fill='both')

# Main tab
main_frame = ttk.Frame(notebook)
notebook.add(main_frame, text='Main')

job_number_label = tk.Label(main_frame, text="Job number:")
job_number_label.pack()

job_number_entry = tk.Entry(main_frame)
job_number_entry.pack()

select_button = tk.Button(main_frame, text="Select directory", command=select_directory)
select_button.pack()

# Configuration tab
config_frame = ttk.Frame(notebook)
notebook.add(config_frame, text='Configuration')

excel_file_label = tk.Label(config_frame, text="Excel file:")
excel_file_label.pack()

excel_file_entry = tk.Entry(config_frame)
excel_file_entry.pack()

select_excel_button = tk.Button(config_frame, text="Select Excel file", command=select_excel_file)
select_excel_button.pack()

# Load the configuration on startup
load_configuration()

root.mainloop()
